"""
Módulo de procesamiento: descarga imágenes, aplica OCR multi-transformación
y calcula coincidencias con los datos del Excel con matching aproximado.

Pipeline de imagen (adaptado del enfoque PHP con Intervention Image):
  - 9 variantes de preprocesamiento x 6 modos PSM de Tesseract
  - Corrección automática de rotación (deskew)
  - Parada anticipada cuando se obtiene texto de buena calidad
  - Fuzzy matching con rapidfuzz para tolerar errores de OCR
  - Generador SSE para streaming fila a fila al frontend
"""
from __future__ import annotations

import io
import os
import re
import time
import unicodedata
import uuid
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterator

import cv2
import numpy as np
import openpyxl
import pypdfium2 as pdfium
import pytesseract
import requests
from PIL import Image
from pypdf import PdfReader
from rapidfuzz import fuzz

# ── Configuración ──────────────────────────────────────────────────────────────

# Orden de búsqueda de Tesseract:
# 1. Variable de entorno TESSERACT_CMD, si el usuario la define.
# 2. Rutas comunes de macOS (incluida tu instalación con Homebrew).
# 3. Rutas comunes de Linux.
# 4. Rutas comunes de Windows.
TESSERACT_CANDIDATES = [
    Path(path)
    for path in [
        os.getenv("TESSERACT_CMD"),
        "/opt/homebrew/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/usr/bin/tesseract",
        "/snap/bin/tesseract",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    if path
]

EXCEL_DATE_ORIGIN = date(1899, 12, 30)

# Transformaciones de imagen ordenadas de mayor a menor probabilidad de éxito
# para cédulas colombianas (equivalentes al PHP con Intervention Image + GD)
TRANSFORM_NAMES: list[str] = [
    "grayscale_only",     # prioridad: gris directo como referencia original
    "soft_gray_clean",    # mejora natural sin posterización
    "light_text",         # normalización + CLAHE + filtro + escalado (claras)
    "pdf_bw_clean",       # fallback
    "scratch_clean",      # fallback
    "tutorial_4_steps", # fallback fuerte cuando lo anterior no funciona
    "binarized_base",     # escala de grises → OTSU  (base más confiable)
    "high_contrast",      # grises → CLAHE alto → OTSU
    "resize_2x",          # grises → OTSU → escalar ×2
    "extra_threshold",    # threshold adaptativo gaussiano
    "fecha_optimizado",   # escalar ×3 → nitidez → OTSU  (optimizado para fechas)
    "sharpen",            # OTSU → unsharp mask
    "brightness",         # normalizar brillo → OTSU
    "fecha_erosion",      # escalar ×4 → OTSU → nitidez (erosión sintética)
]

# PSMs de Tesseract a probar por cada transformación
PSM_MODES: list[int] = [6, 11, 4, 3, 7, 8]

# Umbral de similitud (0–100) para fuzzy matching de palabras
FUZZY_WORD_THRESHOLD = 82

# Texto mínimo para parada anticipada (equivalente al break 2 del PHP)
EARLY_STOP_CHARS = 100
EARLY_STOP_WORDS = 10
BOX_MIN_CONF = 8.0
BOX_MIN_TEXT_LEN = 2
DATE_REGEX_DDMMYYYY = r"\b\d{2}-\d{2}-\d{4}\b"
BOX_PREVIEW_CONF = 25.0
FORCE_UNIFORM_PIPELINE = True

PDF_BASE_URL = "https://dashboard.uttalento.co/dashboard/cedulas"
PDF_FETCH_RETRIES = 4
LOCAL_PDF_DIR = Path(__file__).resolve().parent / "uploads" / "pdf_cache"


def _pdf_request_params() -> dict[str, str]:
    now = int(time.time())
    return {
        "v": str(now),
        "cache": str(now * 1000),
    }


def _pdf_request_headers() -> dict[str, str]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "application/pdf,application/octet-stream;q=0.9,*/*;q=0.8",
        "Referer": f"{PDF_BASE_URL}/",
        "Origin": "https://dashboard.uttalento.co",
    }
    # Si el portal protege los PDFs con sesión, permite inyectar cookie desde entorno.
    cookie = os.getenv("UTTALENTO_COOKIE", "").strip()
    if cookie:
        headers["Cookie"] = cookie
    return headers


def _download_pdf_with_retries(pdf_url: str) -> tuple[bytes | None, str | None]:
    """
    Descarga PDF con varios intentos para evitar falsos negativos por bloqueos transitorios.
    Retorna (pdf_bytes, error_message).
    """
    session = requests.Session()
    last_error = ""

    for attempt in range(1, PDF_FETCH_RETRIES + 1):
        try:
            resp = session.get(
                pdf_url,
                params=_pdf_request_params(),
                headers=_pdf_request_headers(),
                timeout=35,
                allow_redirects=True,
            )
        except Exception as exc:
            last_error = f"Error de red en intento {attempt}: {exc}"
            if attempt < PDF_FETCH_RETRIES:
                time.sleep(1.2 * attempt)
            continue

        if resp.status_code == 404:
            return None, f"No existe PDF para esta cédula ({pdf_url})"

        content_type = (resp.headers.get("content-type") or "").lower()
        is_pdf = "application/pdf" in content_type or resp.content.startswith(b"%PDF")

        if resp.status_code == 200 and is_pdf:
            return resp.content, None

        cf_mitigated = (resp.headers.get("cf-mitigated") or "").lower()
        if resp.status_code == 403 and "challenge" in cf_mitigated:
            last_error = (
                "Cloudflare challenge detectado al descargar el PDF "
                f"(intento {attempt}/{PDF_FETCH_RETRIES})"
            )
        else:
            last_error = (
                f"Respuesta inesperada {resp.status_code} en intento {attempt} "
                f"(content-type: {content_type or 'desconocido'})"
            )

        if attempt < PDF_FETCH_RETRIES:
            time.sleep(1.2 * attempt)

    return None, (
        "No fue posible descargar el PDF para OCR tras varios intentos. "
        "El enlace puede abrir en navegador, pero el backend está recibiendo bloqueo/bot challenge. "
        f"URL: {pdf_url}. Último detalle: {last_error}"
    )


# ── Tesseract ──────────────────────────────────────────────────────────────────

def configure_tesseract() -> None:
    for candidate in TESSERACT_CANDIDATES:
        if candidate.exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return
    raise FileNotFoundError(
        "No se encontró Tesseract en las rutas configuradas. "
        "Si ya está instalado, define la variable de entorno TESSERACT_CMD "
        "con la ruta completa al ejecutable."
    )


def _available_language() -> str:
    try:
        languages = set(pytesseract.get_languages(config=""))
    except pytesseract.TesseractError:
        return "eng"
    return "spa" if "spa" in languages else "eng"


# ── Preprocesamiento de imagen ─────────────────────────────────────────────────

def _otsu(gray: np.ndarray) -> np.ndarray:
    _, t = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return t


def _unsharp(img: np.ndarray, sigma: float = 1.0, strength: float = 1.5) -> np.ndarray:
    """Unsharp mask — aumenta nitidez conservando bordes."""
    blurred = cv2.GaussianBlur(img, (0, 0), sigma)
    return cv2.addWeighted(img, 1.0 + strength, blurred, -strength, 0)


def _gamma_correct(gray: np.ndarray, gamma: float) -> np.ndarray:
    """Ajusta gamma para rescatar texto en imágenes muy claras."""
    safe_gamma = max(gamma, 0.1)
    inv_gamma = 1.0 / safe_gamma
    table = np.array([
        ((i / 255.0) ** inv_gamma) * 255 for i in np.arange(256)
    ]).astype("uint8")
    return cv2.LUT(gray, table)


def _safe_resize(gray: np.ndarray, factor: float, max_pixels: int = 4_000_000) -> np.ndarray:
    """Escala la imagen limitando el resultado a max_pixels para no saturar memoria."""
    h, w = gray.shape[:2]
    effective = min(factor, (max_pixels / (h * w)) ** 0.5)
    if effective <= 1.05:
        return gray
    return cv2.resize(gray, None, fx=effective, fy=effective,
                      interpolation=cv2.INTER_CUBIC)


def _prepare_pdf_page_for_ocr(img_bgr: np.ndarray) -> np.ndarray:
    """
    Prepara una página PDF en blanco y negro para mejorar legibilidad OCR.
    Mantiene bordes y reduce ruido de fondo típico en escaneos.
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)


def _soft_gray_clean(gray: np.ndarray) -> np.ndarray:
    """Preprocesado suave para OCR, prioriza legibilidad sin caricaturizar."""
    base = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
    base = cv2.bilateralFilter(base, 7, 55, 55)
    clahe = cv2.createCLAHE(clipLimit=2.2, tileGridSize=(8, 8))
    contrasted = clahe.apply(base)
    de_scratched = _remove_scratch_lines(contrasted)
    return _unsharp(de_scratched, sigma=0.7, strength=0.9)


def _natural_document_preview(img_bgr: np.ndarray) -> np.ndarray:
    """
    Devuelve la imagen tal como viene del PDF renderizado,
    sin alteraciones visuales para que coincida con el original.
    """
    return img_bgr.copy()


def _trim_whitespace_margins(img_bgr: np.ndarray, margin: int = 14) -> np.ndarray:
    """Recorta márgenes blancos amplios para concentrar OCR en la cédula.

    La función es conservadora: si el recorte resultante es demasiado pequeño o
    no aporta señal clara, devuelve la imagen original.
    """
    if img_bgr is None or img_bgr.size == 0 or len(img_bgr.shape) != 3:
        return img_bgr

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    content_mask = (
        (gray < 246)
        | (hsv[:, :, 1] > 18)
        | (hsv[:, :, 2] < 245)
    ).astype(np.uint8) * 255

    if cv2.countNonZero(content_mask) == 0:
        return img_bgr

    close_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (19, 19))
    content_mask = cv2.morphologyEx(content_mask, cv2.MORPH_CLOSE, close_kernel, iterations=1)
    content_mask = cv2.dilate(content_mask, cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7)), iterations=1)

    ys, xs = np.where(content_mask > 0)
    if ys.size == 0 or xs.size == 0:
        return img_bgr

    y1, y2 = int(ys.min()), int(ys.max())
    x1, x2 = int(xs.min()), int(xs.max())

    h, w = img_bgr.shape[:2]
    pad = max(margin, int(min(h, w) * 0.01))
    y1 = max(0, y1 - pad)
    x1 = max(0, x1 - pad)
    y2 = min(h - 1, y2 + pad)
    x2 = min(w - 1, x2 + pad)

    cropped = img_bgr[y1:y2 + 1, x1:x2 + 1]
    if cropped.size == 0:
        return img_bgr
    if cropped.shape[0] < 120 or cropped.shape[1] < 120:
        return img_bgr
    if (cropped.shape[0] * cropped.shape[1]) < (img_bgr.shape[0] * img_bgr.shape[1] * 0.15):
        return cropped

    return cropped


def _remove_scratch_lines(gray: np.ndarray) -> np.ndarray:
    """
    Reduce rayones largos (horizontales/verticales) sin borrar trazos de letras.
    Mejora OCR cuando el documento tiene marcas sobre el texto.
    """
    if gray is None or gray.size == 0:
        return gray

    # Binario inverso para detectar trazos oscuros.
    _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    inv = 255 - bw

    h_len = max(gray.shape[1] // 28, 28)
    v_len = max(gray.shape[0] // 28, 28)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (h_len, 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_len))

    horizontal = cv2.morphologyEx(inv, cv2.MORPH_OPEN, h_kernel)
    vertical = cv2.morphologyEx(inv, cv2.MORPH_OPEN, v_kernel)

    mask = cv2.bitwise_or(horizontal, vertical)
    if cv2.countNonZero(mask) == 0:
        return gray

    mask = cv2.dilate(mask, np.ones((2, 2), np.uint8), iterations=1)
    cleaned = cv2.inpaint(gray, mask, 2, cv2.INPAINT_TELEA)
    return cleaned


def _tutorial_four_step_preprocess(gray: np.ndarray) -> np.ndarray:
    """
    Pipeline de 4 pasos para cédulas (según guía OCR):
    1) Escala de grises (ya recibida en `gray`)
    2) Binarización OTSU
    3) Apertura morfológica para limpiar ruido fino
    4) Inversión para dejar fondo claro y texto oscuro
    """
    base = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
    de_scratched = _remove_scratch_lines(base)
    smooth = cv2.GaussianBlur(de_scratched, (3, 3), 0)
    otsu = _otsu(smooth)

    # Apertura pequeña para ruido fino y cierre pequeño para recuperar trazos.
    open_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
    close_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    opened = cv2.morphologyEx(otsu, cv2.MORPH_OPEN, open_kernel, iterations=1)
    healed = cv2.morphologyEx(opened, cv2.MORPH_CLOSE, close_kernel, iterations=1)

    # Resultado final del tutorial: fondo blanco con texto oscuro.
    return _normalize_binary_polarity(healed)


def _normalize_binary_polarity(binary_img: np.ndarray) -> np.ndarray:
    """Asegura texto oscuro sobre fondo claro para estabilizar OCR."""
    if binary_img is None or binary_img.size == 0:
        return binary_img

    white_ratio = float(np.mean(binary_img > 127))
    # Si predomina fondo oscuro, invertimos.
    if white_ratio < 0.5:
        return cv2.bitwise_not(binary_img)
    return binary_img


def _collect_text_boxes(processed: np.ndarray, lang: str, psm: int = 11) -> list[tuple[int, int, int, int]]:
    """Obtiene cajas de texto detectadas por Tesseract para aislar regiones útiles."""
    cfg = f"--oem 3 --psm {psm} -c preserve_interword_spaces=1"
    try:
        data = pytesseract.image_to_data(
            processed,
            lang=lang,
            config=cfg,
            output_type=pytesseract.Output.DICT,
        )
    except Exception:
        return []

    boxes: list[tuple[int, int, int, int]] = []
    n = len(data.get("text", []))
    for i in range(n):
        raw_text = (data.get("text", [""])[i] or "").strip()
        cleaned = re.sub(r"[^A-Za-z0-9/\-]", "", raw_text)
        try:
            conf = float(data.get("conf", ["-1"])[i])
        except (TypeError, ValueError):
            continue

        if len(cleaned) < BOX_MIN_TEXT_LEN or conf < BOX_MIN_CONF:
            continue

        x = int(data.get("left", [0])[i])
        y = int(data.get("top", [0])[i])
        w = int(data.get("width", [0])[i])
        h = int(data.get("height", [0])[i])
        if w <= 0 or h <= 0:
            continue
        boxes.append((x, y, w, h))

    return boxes


def _apply_text_box_mask(processed: np.ndarray, boxes: list[tuple[int, int, int, int]], padding: int = 3) -> np.ndarray:
    """Mantiene solo zonas con texto detectado y limpia el resto del fondo."""
    h, w = processed.shape[:2]
    if not boxes:
        return processed

    mask = np.zeros((h, w), dtype=np.uint8)
    for x, y, bw, bh in boxes:
        x1 = max(0, x - padding)
        y1 = max(0, y - padding)
        x2 = min(w, x + bw + padding)
        y2 = min(h, y + bh + padding)
        mask[y1:y2, x1:x2] = 255

    white_bg = np.full_like(processed, 255)
    masked = np.where(mask == 255, processed, white_bg)
    return masked.astype(np.uint8)


def _tutorial_text_box_guided_image(img_bgr: np.ndarray, lang: str) -> np.ndarray:
    """
    Replica el enfoque de 'dibujando cajas':
    1) preprocesa en binario limpio,
    2) detecta cajas de texto con Tesseract,
    3) conserva solo esas regiones para OCR final.
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    base = _tutorial_four_step_preprocess(gray)
    boxes = _collect_text_boxes(base, lang=lang, psm=11)
    guided = _apply_text_box_mask(base, boxes, padding=3)

    # Cierre suave final para mantener letras continuas.
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    guided = cv2.morphologyEx(guided, cv2.MORPH_CLOSE, kernel, iterations=1)
    return _normalize_binary_polarity(guided)


def _rotate_bound(image: np.ndarray, angle: float) -> np.ndarray:
    """Rota la imagen preservando todo el contenido visible."""
    if abs(angle) < 0.01:
        return image

    h, w = image.shape[:2]
    center = (w / 2, h / 2)
    matrix = cv2.getRotationMatrix2D(center, -angle, 1.0)
    cos = abs(matrix[0, 0])
    sin = abs(matrix[0, 1])

    new_w = int((h * sin) + (w * cos))
    new_h = int((h * cos) + (w * sin))

    matrix[0, 2] += (new_w / 2) - center[0]
    matrix[1, 2] += (new_h / 2) - center[1]

    return cv2.warpAffine(
        image,
        matrix,
        (new_w, new_h),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE,
    )


def _detect_skew_angle(gray: np.ndarray) -> float:
    """
    Detecta inclinación fina del documento usando minAreaRect.
    Solo devuelve ángulos entre 0.5° y 20° para evitar falsos positivos.
    """
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    coords = np.column_stack(np.where(binary > 0))
    if len(coords) < 50:
        return 0.0
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
    if abs(angle) < 0.5 or abs(angle) > 20:
        return 0.0
    return angle


def _score_ocr_text(text: str) -> float:
    """
    Puntúa texto OCR favoreciendo palabras reales y etiquetas típicas del documento.
    Evita elegir la salida más larga cuando está llena de ruido.
    """
    normalized = _normalize(text)
    tokens = [token for token in normalized.split() if token]
    if not tokens:
        return 0.0

    long_tokens = [token for token in tokens if len(token) >= 3]
    digit_tokens = [token for token in tokens if token.isdigit() and len(token) >= 4]
    alpha_tokens = [token for token in long_tokens if token.isalpha()]
    useful_chars = sum(len(token) for token in long_tokens)

    keywords = {
        "republica", "colombia", "cedula", "ciudadania", "identificacion",
        "nombres", "apellidos", "sexo", "nacimiento", "fecha",
        "documento", "firma", "numero", "nombre",
    }
    keyword_hits = sum(1 for token in long_tokens if token in keywords)

    repeated_noise = sum(1 for token in long_tokens if len(set(token)) <= 2)
    short_noise = sum(1 for token in tokens if len(token) == 1)

    return (
        useful_chars * 1.2
        + len(alpha_tokens) * 6.0
        + len(digit_tokens) * 4.0
        + keyword_hits * 18.0
        - repeated_noise * 8.0
        - short_noise * 1.5
    )


def _detect_osd_rotation(gray: np.ndarray, lang: str | None = None) -> float:
    """Obtiene el giro grueso sugerido por Tesseract OSD si está disponible."""
    try:
        osd = pytesseract.image_to_osd(gray, lang="osd", config="--psm 0")
    except Exception:
        return 0.0

    match = re.search(r"Rotate:\s+(\d+)", osd)
    if not match:
        return 0.0

    rotation = int(match.group(1)) % 360
    return float(rotation) if rotation in {90, 180, 270} else 0.0


def _osd_confidence(gray: np.ndarray) -> float:
    try:
        osd = pytesseract.image_to_osd(gray, lang="osd", config="--psm 0")
    except Exception:
        return 0.0

    match = re.search(r"Orientation confidence:\s+([0-9.]+)", osd)
    if not match:
        return 0.0
    try:
        return float(match.group(1))
    except ValueError:
        return 0.0


def _orientation_score(img_bgr: np.ndarray, lang: str, osd_rotation: float) -> tuple[float, float]:
    """
    Puntúa una orientación según cuánto texto confiable detecta Tesseract.
    Devuelve (score, suggested_rotation_for_this_view).
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    preview = _apply_transform(img_bgr, "high_contrast")

    words_score = 0.0
    try:
        data = pytesseract.image_to_data(
            preview,
            lang=lang,
            config="--oem 3 --psm 6",
            output_type=pytesseract.Output.DICT,
        )
        confidences: list[float] = []
        text_hits = 0
        for text, conf in zip(data.get("text", []), data.get("conf", [])):
            cleaned = re.sub(r"[^A-Za-z0-9]", "", text or "")
            try:
                conf_value = float(conf)
            except (TypeError, ValueError):
                continue
            if cleaned:
                text_hits += len(cleaned)
            if len(cleaned) >= 3 and conf_value > 0:
                confidences.append(conf_value)

        if confidences:
            words_score += (sum(confidences) / len(confidences)) * 2.0
        words_score += min(text_hits, 120)
    except Exception:
        pass

    coarse_bonus = 0.0
    suggested_rotation = _detect_osd_rotation(gray, lang)
    osd_conf = _osd_confidence(gray)
    if suggested_rotation == 0.0:
        coarse_bonus += 80.0 + (osd_conf * 10.0)
    elif suggested_rotation == osd_rotation:
        coarse_bonus += 20.0 + (osd_conf * 5.0)
    elif suggested_rotation:
        coarse_bonus -= 10.0

    horizontal_bonus = 6.0 if img_bgr.shape[1] >= img_bgr.shape[0] else 0.0
    return words_score + coarse_bonus + horizontal_bonus, suggested_rotation


def _select_best_orientation(img_bgr: np.ndarray, lang: str) -> np.ndarray:
    """Prueba giros cardinales y elige el que deja el texto más legible."""
    candidates = [0.0, 90.0, 180.0, 270.0]
    best_image = img_bgr
    best_score = float("-inf")

    for angle in candidates:
        rotated = _rotate_bound(img_bgr, angle)
        score, _ = _orientation_score(rotated, lang, angle)
        if score > best_score:
            best_score = score
            best_image = rotated

    return best_image


def _orient_image(img_bgr: np.ndarray, lang: str) -> np.ndarray:
    """
    Endereza la imagen en dos fases:
    1. Prueba giros de 0/90/180/270° y elige el que produce texto más legible.
    2. Ajusta la inclinación fina para que el texto quede recto.
    """
    oriented = _select_best_orientation(img_bgr, lang)

    gray = cv2.cvtColor(oriented, cv2.COLOR_BGR2GRAY)
    skew_angle = _detect_skew_angle(gray)
    if skew_angle:
        oriented = _rotate_bound(oriented, skew_angle)

    return oriented


def _apply_transform(img_bgr: np.ndarray, name: str) -> np.ndarray:
    """
    Aplica una de las 9 variantes de preprocesamiento sobre la imagen BGR.
    Equivalente a las funciones crearImagenTemporalIntervention del PHP.
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    if name == "soft_gray_clean":
        return _soft_gray_clean(gray)

    if name == "tutorial_4_steps":
        return _normalize_binary_polarity(_tutorial_four_step_preprocess(gray))

    if name == "pdf_bw_clean":
        prepared = _prepare_pdf_page_for_ocr(img_bgr)
        return cv2.cvtColor(prepared, cv2.COLOR_BGR2GRAY)

    if name == "scratch_clean":
        cleaned = _remove_scratch_lines(gray)
        sharpened = _unsharp(cleaned, sigma=0.7, strength=1.4)
        out = cv2.adaptiveThreshold(
            sharpened,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            31,
            10,
        )
        return _normalize_binary_polarity(out)

    if name == "light_text":
        # Refuerzo para imágenes claras: recupera contraste local y sube resolución útil.
        normalized = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
        gamma = _gamma_correct(normalized, 1.6)
        clahe = cv2.createCLAHE(clipLimit=3.5, tileGridSize=(8, 8))
        contrasted = clahe.apply(gamma)
        filtered = cv2.bilateralFilter(contrasted, 9, 75, 75)
        scaled = _safe_resize(filtered, 2.8)
        sharpened = _unsharp(scaled, sigma=0.8, strength=1.8)
        return cv2.adaptiveThreshold(
            sharpened,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            31,
            9,
        )

    if name == "grayscale_only":
        # PASO 1 del PHP: solo escala de grises
        return gray

    if name == "binarized_base":
        # PASO 1+2 del PHP: grises → binarización OTSU
        return _otsu(gray)

    if name == "high_contrast":
        # PASO 1+2+contraste_alto: CLAHE potente antes de OTSU
        clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
        return _otsu(clahe.apply(gray))

    if name == "resize_2x":
        # PASO 1+2+redimensionar: escalar ×2 (PHP resize w*2, h*2)
        return _otsu(_safe_resize(gray, 2.0))

    if name == "sharpen":
        # PASO 1+2+nitidez: unsharp mask sobre imagen binarizada
        return _unsharp(_otsu(gray))

    if name == "brightness":
        # PASO 1+2+brillo: normalizar rango dinámico → OTSU
        normalized = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
        return _otsu(normalized)

    if name == "extra_threshold":
        # PASO 1+2+umbral_extra: threshold adaptativo gaussiano
        return cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 15, 10
        )

    if name == "fecha_optimizado":
        # PHP fechas_optimizado: ×3 zoom + nitidez alta + contraste alto
        scaled = _safe_resize(gray, 3.0)
        sharpened = _unsharp(scaled, sigma=1.0, strength=2.0)
        return _otsu(sharpened)

    if name == "fecha_erosion":
        # PHP fechas_erosion: ×4 zoom + OTSU + nitidez fuerte
        scaled = _safe_resize(gray, 4.0)
        binarized = _otsu(scaled)
        return _unsharp(binarized, sigma=0.5, strength=1.5)

    return gray


def _ocr_text_from_data(processed: np.ndarray, lang: str, psm: int) -> tuple[str, float]:
    """
    Ejecuta OCR usando image_to_data para filtrar ruido por confianza y longitud.
    Devuelve (texto, score_confianza).
    """
    cfg = (
        f"--oem 3 --psm {psm} "
        "-c preserve_interword_spaces=1 "
        "-c textord_heavy_nr=1"
    )

    try:
        data = pytesseract.image_to_data(
            processed,
            lang=lang,
            config=cfg,
            output_type=pytesseract.Output.DICT,
        )
    except Exception:
        return "", 0.0

    words: list[str] = []
    confs: list[float] = []
    for raw_text, raw_conf in zip(data.get("text", []), data.get("conf", [])):
        text = (raw_text or "").strip()
        if not text:
            continue

        try:
            conf = float(raw_conf)
        except (TypeError, ValueError):
            continue

        # Ignora ruido de baja confianza y restos de rayones.
        cleaned = re.sub(r"[^A-Za-z0-9/\-]", "", text)
        is_short_digit = len(cleaned) == 1 and cleaned.isdigit()
        if (len(cleaned) < 2 and not is_short_digit) or conf < 12:
            continue

        words.append(cleaned)
        confs.append(conf)

    if not words:
        return "", 0.0

    text_joined = " ".join(words)
    avg_conf = sum(confs) / len(confs)
    return text_joined, avg_conf


def _boost_black_text(binary_like: np.ndarray) -> np.ndarray:
    """Refuerza trazos negros (nombres/letras) en imagen B/N para mejorar OCR."""
    if len(binary_like.shape) == 3:
        gray = cv2.cvtColor(binary_like, cv2.COLOR_BGR2GRAY)
    else:
        gray = binary_like

    # Asegura imagen binaria estable.
    bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

    # Texto negro sobre fondo blanco -> invertir para engrosar texto.
    text_mask = 255 - bw
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    text_mask = cv2.dilate(text_mask, kernel, iterations=1)
    text_mask = cv2.morphologyEx(text_mask, cv2.MORPH_CLOSE, kernel, iterations=1)
    boosted = 255 - text_mask
    return boosted


def _ocr_tutorial(image: np.ndarray, lang: str) -> str:
    """OCR en español probando psm 11/6/4 y eligiendo el texto con mejor score."""
    ocr_lang = "spa" if lang == "spa" else lang
    variants = [image, _boost_black_text(image)]
    best_text = ""
    best_score = float("-inf")

    for variant in variants:
        for psm in (11, 6, 4):
            config = f"--oem 3 -l {ocr_lang} --psm {psm} -c preserve_interword_spaces=1"
            try:
                text = pytesseract.image_to_string(variant, config=config).strip()
            except Exception:
                continue
            if not text:
                continue

            score = _score_ocr_text(text)
            if score > best_score or (score == best_score and len(text) > len(best_text)):
                best_text = text
                best_score = score

    return best_text


def _robust_fallback_ocr(img_bgr: np.ndarray, lang: str) -> str:
    """OCR de respaldo para evitar salidas vacías cuando falla el flujo principal."""
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(gray)
    otsu = cv2.threshold(clahe, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    inv = 255 - cv2.threshold(clahe, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

    variants = [gray, clahe, otsu, inv]
    best_text = ""
    best_score = float("-inf")

    for variant in variants:
        for psm in (6, 11, 3, 4):
            try:
                text = pytesseract.image_to_string(
                    variant,
                    lang=lang,
                    config=f"--oem 3 --psm {psm} -c preserve_interword_spaces=1",
                ).strip()
            except Exception:
                continue

            if not text:
                continue

            score = _score_ocr_text(text)
            if score > best_score or (score == best_score and len(text) > len(best_text)):
                best_text = text
                best_score = score

    return best_text


_MONTHS_RE = re.compile(
    r"\b(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\b", re.IGNORECASE
)
_CEDULA_NUM_RE = re.compile(r"\d[\d.,]{6,}")


def _score_cedula_front(text: str) -> float:
    """Score especializado para texto del frontal de una cédula colombiana.

    Premia fuertemente patrones como número de cédula, fechas con mes abreviado
    y palabras clave del documento; penaliza texto que parece MRZ o texto
    invertido (letras/números mezclados sin separación).
    """
    norm = _normalize(text)
    tokens = [t for t in norm.split() if t]
    if not tokens:
        return 0.0

    alpha_words = [t for t in tokens if t.isalpha() and len(t) >= 4]
    long_alpha = [t for t in tokens if t.isalpha() and len(t) >= 6]
    mixed_noise = sum(
        1 for t in tokens if len(t) >= 5 and not t.isalpha() and not t.isdigit()
    )
    cedula_num = len(_CEDULA_NUM_RE.findall(norm))
    month_hits = len(_MONTHS_RE.findall(norm))
    keywords = {"republica", "colombia", "cedula", "ciudadania", "identificacion",
                "nombres", "apellidos", "nacimiento", "fecha", "numero"}
    keyword_hits = sum(1 for t in tokens if t in keywords)

    return (
        len(alpha_words) * 8.0
        + len(long_alpha) * 6.0
        + cedula_num * 60.0
        + month_hits * 50.0
        + keyword_hits * 25.0
        - mixed_noise * 5.0
    )


def _front_name_focused_ocr(img_bgr: np.ndarray, lang: str) -> str:
    """OCR especializado para frontal: realza texto oscuro (nombres/apellidos).

    Usa dos kernels blackhat (25 y 15) para capturar tanto números grandes
    como texto fino de nombres/apellidos en cédulas con distinto diseño.
    Cuando recibe una página completa (retrato, alto > ancho), prueba la
    mitad superior Y la mitad inferior (por si la imagen fue rotada 180°)
    y conserva la que produce mejor texto.
    """
    h_full, w_full = img_bgr.shape[:2]
    close_k = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))

    def _blackhat_best(crop: np.ndarray) -> tuple[str, float]:
        """Aplica blackhat con kernels 25 y 15 y devuelve (texto, score cedula)."""
        h, w = crop.shape[:2]
        focus_crops: list[np.ndarray] = [crop]

        # En documentos horizontales, la información nominal suele estar
        # concentrada en la zona izquierda (número, apellidos, nombres).
        if w > h * 1.1:
            x1 = int(w * 0.04)
            x2 = int(w * 0.72)
            # Franja superior izquierda: suele contener número, apellidos y nombres.
            focus_crops.append(crop[int(h * 0.18):int(h * 0.48), int(w * 0.04):int(w * 0.52)])
            focus_crops.append(crop[int(h * 0.24):int(h * 0.42), int(w * 0.04):int(w * 0.46)])
            focus_crops.append(crop[:, x1:x2])
            focus_crops.append(crop[int(h * 0.08):int(h * 0.94), x1:x2])
        else:
            focus_crops.append(crop[int(h * 0.06):int(h * 0.94), int(w * 0.06):int(w * 0.94)])

        best_text = ""
        best_score = float("-inf")
        for focus in focus_crops:
            if focus.size == 0:
                continue
            gray = cv2.cvtColor(focus, cv2.COLOR_BGR2GRAY)
            for ksz in (25, 15):
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (ksz, ksz))
                bh = cv2.morphologyEx(gray, cv2.MORPH_BLACKHAT, kernel)
                bh = cv2.normalize(bh, None, 0, 255, cv2.NORM_MINMAX)
                thr = cv2.threshold(bh, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
                thr = cv2.morphologyEx(thr, cv2.MORPH_CLOSE, close_k, iterations=1)
                inv = 255 - thr
                up = cv2.resize(inv, None, fx=2.5, fy=2.5, interpolation=cv2.INTER_CUBIC)
                for psm in (6, 4):
                    try:
                        text = pytesseract.image_to_string(
                            up, lang=lang,
                            config=f"--oem 1 --psm {psm} -c preserve_interword_spaces=1",
                            timeout=6,
                        ).strip()
                    except Exception:
                        continue
                    if not text:
                        continue
                    score = _score_cedula_front(text)
                    if score > best_score or (score == best_score and len(text) > len(best_text)):
                        best_text = text
                        best_score = score
        return best_text, best_score

    # Para imágenes de página completa (retrato), prueba ambas mitades porque
    # _orient_image puede haber rotado la página 180° poniendo el frontal abajo.
    if h_full > w_full * 1.2:
        top_crop = img_bgr[int(h_full * 0.04): int(h_full * 0.58), :]
        bot_crop = img_bgr[int(h_full * 0.42): int(h_full * 0.96), :]
        # Candidatos: mitad superior normal, mitad inferior normal y
        # mitad inferior rotada 180° (por si la página fue girada).
        bot_rot = cv2.rotate(bot_crop, cv2.ROTATE_180)
        candidates_page = [
            _blackhat_best(top_crop),
            _blackhat_best(bot_crop),
            _blackhat_best(bot_rot),
        ]
        best_text, _ = max(candidates_page, key=lambda x: x[1])
        return best_text

    # Para imágenes ya recortadas (región de cédula), aplicar directamente.
    text, _ = _blackhat_best(img_bgr)
    return text


def _is_useful_ocr_text(text: str, max_len: int = 1200) -> bool:
    """Descarta salidas OCR claramente ruidosas para no contaminar el matching."""
    if not text:
        return False
    raw = text.strip()
    if len(raw) < 18 or len(raw) > max_len:
        return False

    norm = _normalize(raw)
    tokens = [t for t in norm.split() if t]
    if len(tokens) < 3:
        return False

    letters = sum(1 for ch in norm if ch.isalpha())
    digits = sum(1 for ch in norm if ch.isdigit())
    density = (letters + digits) / max(len(norm), 1)
    if density < 0.28:
        return False

    has_keyword = any(k in norm for k in (
        "cedula", "ciudadania", "nombres", "apellidos", "fecha", "expedicion", "colombia"
    ))
    has_long_digits = bool(re.search(r"\d{7,}", norm))
    has_name_like = len([t for t in tokens if t.isalpha() and len(t) >= 4]) >= 2

    return has_keyword or has_long_digits or has_name_like


def _should_mark_invalid_clarity(ocr_text: str, match_result: dict | None) -> bool:
    """Determina si el OCR debe marcarse como no válido por baja claridad."""
    if not ocr_text or not ocr_text.strip():
        return True

    match_result = match_result or {}
    pct = int(match_result.get("porcentaje", 0) or 0)
    if pct >= 60:
        return False

    norm = _normalize(ocr_text)
    tokens = [t for t in norm.split() if t]
    if len(tokens) < 12:
        return True

    letters = sum(1 for ch in norm if ch.isalpha())
    digits = sum(1 for ch in norm if ch.isdigit())
    density = (letters + digits) / max(len(norm), 1)

    keywords = (
        "cedula", "ciudadania", "nombres", "apellidos",
        "nacimiento", "identificacion", "numero", "fecha",
    )
    keyword_hits = sum(1 for k in keywords if k in norm)
    has_doc_like = bool(re.search(r"\d[\d.,]{6,}", ocr_text))
    has_date_like = bool(_extract_date_candidates(ocr_text))

    field_hits = 0
    for v in match_result.values():
        if isinstance(v, dict) and v.get("encontrado"):
            field_hits += 1

    if field_hits <= 2:
        return True
    if density < 0.24 and keyword_hits < 2:
        return True
    if not has_doc_like and not has_date_like and keyword_hits < 3:
        return True

    return False


def _extract_card_regions(img_bgr: np.ndarray) -> list[np.ndarray]:
    """
    Detecta regiones de cédula (baja saturación sobre fondo color) y retorna crops.
    Si no encuentra regiones confiables, retorna la imagen completa.
    """
    h, w = img_bgr.shape[:2]
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    # Cédula: baja saturación + valor medio/alto. Fondo azul: saturación alta.
    mask = cv2.inRange(hsv, (0, 0, 55), (180, 105, 255))
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (9, 9))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=1)

    # Segmentación por bandas horizontales para separar frontal y reverso.
    row_hits = np.count_nonzero(mask > 0, axis=1)
    row_thr = int(w * 0.20)
    active_rows = row_hits > row_thr

    bands: list[tuple[int, int]] = []
    start = None
    for i, active in enumerate(active_rows):
        if active and start is None:
            start = i
        elif not active and start is not None:
            if (i - start) >= int(h * 0.08):
                bands.append((start, i - 1))
            start = None
    if start is not None and (h - start) >= int(h * 0.08):
        bands.append((start, h - 1))

    crops: list[np.ndarray] = []
    for y1, y2 in bands[:3]:
        band = mask[y1:y2 + 1, :]
        col_hits = np.count_nonzero(band > 0, axis=0)
        col_thr = int((y2 - y1 + 1) * 0.12)
        active_cols = np.where(col_hits > col_thr)[0]
        if active_cols.size == 0:
            continue

        x1 = int(active_cols[0])
        x2 = int(active_cols[-1])
        if (x2 - x1) < int(w * 0.25):
            continue

        pad = max(8, int(min(y2 - y1 + 1, x2 - x1 + 1) * 0.03))
        xx1 = max(0, x1 - pad)
        yy1 = max(0, y1 - pad)
        xx2 = min(w - 1, x2 + pad)
        yy2 = min(h - 1, y2 + pad)
        crops.append(img_bgr[yy1:yy2 + 1, xx1:xx2 + 1])

    return crops if crops else [img_bgr]


def _tutorial_main_process_from_image(img_bgr: np.ndarray, lang: str) -> tuple[dict[str, str], np.ndarray, str]:
    """
    Implementa la lógica final solicitada por el usuario (equivalente a main_process):
    - gray -> OTSU_INV -> opening(1,2) -> invert -> OCR
    - si detecta fechas: salida de reverso
    - si no: opening(4,4) -> invert -> OCR y extrae CC/nombres/apellidos
    """
    img_gris = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    thresh = cv2.threshold(img_gris, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 2))
    opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
    invert = 255 - opening
    invert_boosted = _boost_black_text(invert)
    text = _ocr_tutorial(invert_boosted, lang)
    frontal = re.findall(r"\w{2}-\w{2}-\w{4}", text)

    if len(frontal) > 0:
        try:
            output: dict[str, str] = {
                "Fecha de nacimiento": frontal[0],
                "Fecha expedicion": frontal[1],
            }
        except Exception:
            output = {
                "Fecha de nacimiento": "DD-MMM-YYYY",
                "Fecha expedicion": "DD-MMM-YYYY",
            }
        return output, invert, text

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (4, 4))
    opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
    invert = 255 - opening
    invert_boosted = _boost_black_text(invert)
    text = _ocr_tutorial(invert_boosted, lang)
    number = re.findall(r"([0-9])", text)
    number_out = "".join(number)
    number_lines = text.splitlines()
    nombres = number_lines[8].strip() if len(number_lines) > 8 else ""
    apellidos = number_lines[6].strip() if len(number_lines) > 6 else ""
    output = {
        "CC": number_out,
        "Nombres": nombres,
        "Apellidos": apellidos,
    }
    return output, invert, text


def _tutorial_preprocess_steps(img_bgr: np.ndarray, kernel_shape: tuple[int, int]) -> dict[str, np.ndarray]:
    """Pipeline del tutorial: gris -> OTSU_INV -> apertura -> inversión."""
    img_gris = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    thresh = cv2.threshold(img_gris, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, kernel_shape)
    opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
    invert = 255 - opening
    return {
        "gray": img_gris,
        "thresh": thresh,
        "opening": opening,
        "invert": invert,
    }


def _tutorial_ocr_with_boxes(
    processed: np.ndarray,
    lang: str,
    psm: int = 11,
    min_conf: float = BOX_MIN_CONF,
) -> tuple[str, float, list[tuple[int, int, int, int, float, str]]]:
    """Ejecuta OCR por cajas (image_to_data) y devuelve texto + score + cajas válidas."""
    cfg = f"--oem 3 --psm {psm} -c preserve_interword_spaces=1"
    try:
        data = pytesseract.image_to_data(
            processed,
            lang=lang,
            config=cfg,
            output_type=pytesseract.Output.DICT,
        )
    except Exception:
        return "", 0.0, []

    words: list[str] = []
    confs: list[float] = []
    boxes: list[tuple[int, int, int, int, float, str]] = []

    n = len(data.get("text", []))
    for i in range(n):
        raw_text = (data.get("text", [""])[i] or "").strip()
        if not raw_text:
            continue
        cleaned = re.sub(r"[^A-Za-z0-9/\-]", "", raw_text)
        if len(cleaned) < BOX_MIN_TEXT_LEN:
            continue

        try:
            conf = float(data.get("conf", ["-1"])[i])
        except (TypeError, ValueError):
            continue
        if conf < min_conf:
            continue

        x = int(data.get("left", [0])[i])
        y = int(data.get("top", [0])[i])
        w = int(data.get("width", [0])[i])
        h = int(data.get("height", [0])[i])
        if w <= 0 or h <= 0:
            continue

        words.append(cleaned)
        confs.append(conf)
        boxes.append((x, y, w, h, conf, cleaned))

    if not words:
        return "", 0.0, []

    text_joined = " ".join(words)
    avg_conf = sum(confs) / len(confs)
    return text_joined, avg_conf, boxes


def _draw_text_boxes(img_bgr: np.ndarray, boxes: list[tuple[int, int, int, int, float, str]]) -> np.ndarray:
    """Dibuja cajas OCR sobre la imagen para depuración visual."""
    canvas = img_bgr.copy()
    for x, y, w, h, conf, text in boxes:
        cv2.rectangle(canvas, (x, y), (x + w, y + h), (36, 200, 56), 2)
        label = f"{text} ({int(conf)})"
        y_text = max(12, y - 4)
        cv2.putText(canvas, label[:40], (x, y_text), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (36, 200, 56), 1, cv2.LINE_AA)
    return canvas


def _tutorial_link_ocr_and_preview(img_bgr: np.ndarray, lang: str) -> tuple[str, float, np.ndarray]:
    """
    Usa la lógica final tipo main_process solicitada por el usuario, sin cajas.
    """
    output, invert, text = _tutorial_main_process_from_image(img_bgr, lang)
    # Reforzar texto con valores estructurados detectados para mejorar matching.
    structured = " ".join(v for v in output.values() if v)
    text_out = f"{text}\n{structured}".strip()
    _ = invert
    preview_clean = _natural_document_preview(img_bgr)
    return text_out, 0.0, preview_clean


def preprocess_and_ocr(image_path: Path, lang: str) -> tuple[str, np.ndarray]:
    """
    Aplica múltiples transformaciones × varios modos PSM de Tesseract.
    Devuelve el texto más largo encontrado.
    Parada anticipada (equivalente al break 2 del PHP) cuando se detecta
    texto de buena calidad: más de EARLY_STOP_CHARS caracteres y
    EARLY_STOP_WORDS palabras.
    """
    img_bgr = cv2.imread(str(image_path))
    if img_bgr is None:
        raise FileNotFoundError(f"No se pudo abrir la imagen: {image_path}")

    # Mantén una copia exacta para vista: debe coincidir 1:1 con el PDF renderizado.
    preview_original = img_bgr.copy()
    regions = _extract_card_regions(preview_original)
    if len(regions) == 1:
        h, w = preview_original.shape[:2]
        top = preview_original[int(h * 0.16):int(h * 0.60), int(w * 0.07):int(w * 0.93)]
        bottom = preview_original[int(h * 0.50):int(h * 0.96), int(w * 0.07):int(w * 0.93)]
        if top.size and bottom.size:
            regions = [top, bottom]
    # OCR sí puede usar versión orientada para mejorar lectura.
    img_bgr = _orient_image(img_bgr, lang)

    # Modo uniforme: todas las imágenes pasan por exactamente el mismo flujo.
    if FORCE_UNIFORM_PIPELINE:
        try:
            region_texts: list[str] = []
            for idx, region in enumerate(regions):
                # Evalúa orientación original, sugerida por OSD y 180°.
                oriented_auto = _orient_image(region, lang)
                oriented_180 = cv2.rotate(region, cv2.ROTATE_180)
                orient_variants: list[np.ndarray] = [region]
                if not np.array_equal(oriented_auto, region):
                    orient_variants.append(oriented_auto)
                if not any(np.array_equal(oriented_180, v) for v in orient_variants):
                    orient_variants.append(oriented_180)

                # Preselección rápida de orientación para evitar ejecutar el
                # pipeline completo en todas las variantes (acelera mucho).
                ranked_orients: list[tuple[float, np.ndarray]] = []
                for region_oriented in orient_variants:
                    t_probe, _c, _p = _tutorial_link_ocr_and_preview(region_oriented, lang)
                    probe_score = _score_ocr_text(t_probe) + (_score_cedula_front(t_probe) * 0.25)
                    ranked_orients.append((probe_score, region_oriented))

                ranked_orients.sort(key=lambda x: x[0], reverse=True)
                eval_orients = [ranked_orients[0][1]] if ranked_orients else []
                # Si la mejor orientación aún es débil, prueba una segunda.
                if ranked_orients and ranked_orients[0][0] < 90 and len(ranked_orients) > 1:
                    eval_orients.append(ranked_orients[1][1])

                region_candidates: list[str] = []
                front_candidates: list[str] = []
                for region_oriented in eval_orients:
                    t_main, _c, _p = _tutorial_link_ocr_and_preview(region_oriented, lang)

                    # Fallback pesado solo cuando la lectura principal es débil.
                    t_fb = ""
                    if (not _is_useful_ocr_text(t_main)) or (_score_ocr_text(t_main) < 95):
                        t_fb = _robust_fallback_ocr(region_oriented, lang)

                    candidates = [t for t in (t_main, t_fb) if _is_useful_ocr_text(t)]

                    # En cédulas escaneadas la cara frontal puede caer en región 0 o 1.
                    # Ejecutamos OCR frontal en ambas para robustez.
                    if idx <= 1:
                        t_front = _front_name_focused_ocr(region_oriented, lang)
                        if _is_useful_ocr_text(t_front, max_len=1400):
                            candidates.append(t_front)
                            front_candidates.append(t_front)

                    if not candidates:
                        # Si todo parece ruido, usa el menos malo entre main/fallback.
                        fallback_best = t_main if _score_ocr_text(t_main) >= _score_ocr_text(t_fb) else t_fb
                        if fallback_best and fallback_best.strip():
                            candidates = [fallback_best.strip()]

                    if not candidates:
                        continue

                    best_local = max(
                        candidates,
                        key=lambda t: (_score_ocr_text(t) + (_score_cedula_front(t) * 0.35)),
                    )
                    region_candidates.append(best_local.strip())

                if not region_candidates:
                    continue

                best = max(
                    region_candidates,
                    key=lambda t: (_score_ocr_text(t) + (_score_cedula_front(t) * 0.35)),
                )
                region_texts.append(best.strip())

                # Agrega un suplemento frontal para no perder apellidos cuando
                # el mejor score global favorece texto más largo pero menos nominal.
                if front_candidates:
                    best_front = max(front_candidates, key=_score_cedula_front)
                    if _score_cedula_front(best_front) >= 120:
                        region_texts.append(best_front.strip())

            # Quita duplicados y limita longitud total para no contaminar matching.
            uniq: list[str] = []
            seen: set[str] = set()
            for txt in region_texts:
                key = _normalize(txt)
                if key in seen:
                    continue
                seen.add(key)
                uniq.append(txt)

            tutorial_text = "\n".join(uniq)[:1600].strip()
            weak = _score_ocr_text(tutorial_text) < 45 or len(tutorial_text) < 24
            if weak:
                fallback_text = _robust_fallback_ocr(img_bgr, lang)
                if fallback_text:
                    tutorial_text = fallback_text
            return tutorial_text, preview_original
        except Exception:
            fallback = _robust_fallback_ocr(img_bgr, lang)
            return fallback, preview_original

    tutorial_guided: np.ndarray | None = None
    best_preview = img_bgr.copy()

    best_text = ""
    best_score = float("-inf")
    candidate_texts: list[tuple[float, str]] = []

    # Primer intento: ruta exacta del tutorial compartido por el usuario.
    try:
        tutorial_text, tutorial_conf, tutorial_preview = _tutorial_link_ocr_and_preview(img_bgr, lang)
        if tutorial_text:
            tutorial_score = _score_ocr_text(tutorial_text) + 25.0 + (tutorial_conf * 0.6)
            candidate_texts.append((tutorial_score, tutorial_text))
            best_text = tutorial_text
            best_score = tutorial_score
            best_preview = tutorial_preview
            if tutorial_score > 220 and len(tutorial_text) > EARLY_STOP_CHARS:
                merged = _merge_top_ocr_candidates(candidate_texts)
                return merged or best_text, best_preview
    except Exception:
        pass

    for transform_name in TRANSFORM_NAMES:
        try:
            if transform_name == "tutorial_4_steps":
                if tutorial_guided is None:
                    tutorial_guided = _tutorial_text_box_guided_image(img_bgr, lang)
                processed = tutorial_guided
            else:
                processed = _apply_transform(img_bgr, transform_name)
        except Exception:
            continue

        for psm in PSM_MODES:
            text, avg_conf = _ocr_text_from_data(processed, lang, psm)

            if not text:
                try:
                    text = pytesseract.image_to_string(
                        processed,
                        lang=lang,
                        config=f"--oem 3 --psm {psm} -c preserve_interword_spaces=1",
                    ).strip()
                    avg_conf = 0.0
                except Exception:
                    continue

            score = _score_ocr_text(text)
            if avg_conf:
                score += avg_conf * 0.9
            if score > best_score or (score == best_score and len(text) > len(best_text)):
                best_text = text
                best_score = score
                if len(processed.shape) == 2:
                    best_preview = cv2.cvtColor(processed, cv2.COLOR_GRAY2BGR)
                else:
                    best_preview = processed.copy()

            if text:
                candidate_texts.append((score, text))

            # Parada anticipada — texto de calidad suficiente encontrado
            if score > 220 and len(text) > EARLY_STOP_CHARS and len(text.split()) > EARLY_STOP_WORDS:
                candidate_texts.append((score + 1, text))
                merged = _merge_top_ocr_candidates(candidate_texts)
                return merged or best_text, best_preview

    merged = _merge_top_ocr_candidates(candidate_texts)
    return merged or best_text, best_preview


def _merge_top_ocr_candidates(candidates: list[tuple[float, str]], max_items: int = 3) -> str:
    """Combina las mejores lecturas OCR sin repetir salidas casi idénticas."""
    selected: list[str] = []
    selected_norms: list[str] = []

    for _, text in sorted(candidates, key=lambda item: (item[0], len(item[1])), reverse=True):
        normalized = _normalize(text)
        if not normalized:
            continue
        if any(fuzz.ratio(normalized, prev) >= 92 for prev in selected_norms):
            continue
        selected.append(text)
        selected_norms.append(normalized)
        if len(selected) >= max_items:
            break

    return "\n".join(selected)


# ── Normalización de texto ─────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Minúsculas, sin tildes, sin caracteres especiales."""
    text = unicodedata.normalize("NFD", text.lower())
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9/ ]", " ", text)


# ── Fuzzy matching (rapidfuzz) ─────────────────────────────────────────────────

def _fuzzy_word_found(ocr_words: list[str], needle: str) -> bool:
    """
    Busca needle entre las palabras del OCR usando similitud de caracteres.
    Usa rapidfuzz.fuzz.ratio para comparar cada par.
    """
    for word in ocr_words:
        if fuzz.ratio(needle, word) >= FUZZY_WORD_THRESHOLD:
            return True
    return False


def _extract_digit_candidates(text: str, min_len: int = 6, max_len: int = 12) -> set[str]:
    """Extrae secuencias numéricas candidatas desde texto OCR/PDF."""
    raw = re.sub(r"[^0-9]", " ", text)
    return {
        token
        for token in raw.split()
        if min_len <= len(token) <= max_len
    }


def _best_window_similarity(ocr_norm: str, needle_norm: str) -> int:
    """Busca la mejor similitud de needle dentro de ventanas del OCR."""
    ocr_tokens = [t for t in ocr_norm.split() if t]
    needle_tokens = [t for t in needle_norm.split() if t]
    if not ocr_tokens or not needle_tokens:
        return 0

    window = len(needle_tokens)
    best = 0
    for i in range(0, max(len(ocr_tokens) - window + 1, 1)):
        candidate = " ".join(ocr_tokens[i:i + window])
        best = max(best, fuzz.token_set_ratio(needle_norm, candidate))
        if best >= 97:
            return best
    return best


def _check_document_field(ocr_text: str, value: Any) -> tuple[bool, str]:
    """Comparación robusta para número de documento."""
    fmt = str(value) if value is not None else ""
    doc = _clean_doc_number(value)
    if not doc:
        return False, fmt

    candidates = _extract_digit_candidates(ocr_text)
    if doc in candidates:
        return True, fmt

    # Permite detectar el documento cuando viene embebido en secuencias largas
    # (por ejemplo, líneas de código de barras en el reverso de la cédula).
    digit_stream = re.sub(r"\D", "", ocr_text)
    if doc and doc in digit_stream:
        return True, fmt

    # Tolera una confusión OCR leve en dígitos.
    for cand in candidates:
        if abs(len(cand) - len(doc)) <= 1 and fuzz.ratio(cand, doc) >= 92:
            return True, fmt
    return False, fmt


def _check_name_field(ocr_norm: str, value: Any) -> tuple[bool, str]:
    """Comparación robusta para nombres/apellidos."""
    if value is None:
        return False, ""
    needle = _normalize(str(value)).strip()
    if not needle:
        return False, str(value)

    if needle in ocr_norm:
        return True, str(value)

    similarity = _best_window_similarity(ocr_norm, needle)
    if similarity >= 84:
        return True, str(value)

    # Fallback por tokens individuales.
    ocr_words = [w for w in ocr_norm.split() if w]
    needle_words = [w for w in needle.split() if w]
    if needle_words:
        found = all(_fuzzy_word_found(ocr_words, nw) for nw in needle_words)
        return found, str(value)

    return False, str(value)


def _extract_date_candidates(text: str) -> set[str]:
    """Extrae fechas candidatas del texto OCR/PDF en múltiples formatos."""
    candidates: set[str] = set()
    text = re.sub(r"(\d)\s+(\d{3})", r"\1\2", text)

    def _normalize_year(year_txt: str) -> str:
        y = re.sub(r"\D", "", year_txt)
        if len(y) == 4:
            return y
        if len(y) == 3:
            return f"1{y}"
        if len(y) == 2:
            yy = int(y)
            return f"19{y}" if yy >= 30 else f"20{y}"
        return ""

    # dd/mm/yyyy o dd-mm-yyyy
    for day, month, year in re.findall(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", text):
        yy = _normalize_year(year)
        if yy:
            candidates.add(f"{int(day):02d}{int(month):02d}{yy}")

    # ddmmyyyy continuo
    for token in re.findall(r"\b\d{8}\b", text):
        candidates.add(token)

    # dd-mon-yyyy (mes en texto), común en OCR de cédulas reverso.
    month_map = {
        "ene": "01", "jan": "01",
        "feb": "02",
        "mar": "03",
        "abr": "04", "apr": "04",
        "may": "05",
        "jun": "06",
        "jul": "07",
        "ago": "08", "aug": "08",
        "sep": "09", "set": "09",
        "oct": "10",
        "nov": "11",
        "dic": "12", "dec": "12",
    }
    for day, mon_txt, year in re.findall(r"\b(\d{1,2})[\-/ ]([A-Za-z0-9]{3,4})[\-/ ](\d{2,4})\b", text):
        yy = _normalize_year(year)
        if not yy:
            continue

        raw = mon_txt.lower()
        opts = {
            raw,
            raw.replace("0", "o").replace("1", "i").replace("5", "s"),
            raw.replace("0", "o").replace("1", "t").replace("5", "s"),
        }

        mon = None
        for opt in opts:
            mon = month_map.get(opt)
            if mon:
                break

        if not mon:
            best = None
            best_ratio = 0
            for opt in opts:
                for k in month_map:
                    r = fuzz.ratio(opt, k)
                    if r > best_ratio:
                        best_ratio = r
                        best = k
            if best and best_ratio >= 60:
                mon = month_map.get(best)

        if mon:
            candidates.add(f"{int(day):02d}{mon}{yy}")

    return candidates


def _check_field(ocr_norm: str, value: Any) -> tuple[bool, str]:
    """
    Comprueba si el valor del campo aparece en el texto OCR normalizado.
    Primero intenta coincidencia exacta como subcadena, luego fuzzy por palabras.
    """
    if value is None:
        return False, ""
    needle = _normalize(str(value)).strip()
    if not needle:
        return False, str(value)

    # 1. Coincidencia exacta
    if needle in ocr_norm:
        return True, str(value)

    # 2. Fuzzy: todos los tokens del campo deben encontrarse en el OCR
    ocr_words = [w for w in ocr_norm.split() if w]
    needle_words = [w for w in needle.split() if w]
    if needle_words:
        found = all(_fuzzy_word_found(ocr_words, nw) for nw in needle_words)
        return found, str(value)

    return False, str(value)


# ── Fechas ─────────────────────────────────────────────────────────────────────

def _excel_serial_to_date(serial: int) -> date:
    return EXCEL_DATE_ORIGIN + timedelta(days=int(serial))


def _clean_doc_number(value: Any) -> str:
    """Normaliza la cédula para construir el nombre del PDF remoto."""
    if value is None:
        return ""

    # Excel suele entregar cédulas numéricas como float (ej: 51662333.0).
    # En ese caso se debe usar el entero exacto para no romper la URL.
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ""

    raw = str(value).strip()
    if not raw:
        return ""

    try:
        dec = Decimal(raw.replace(",", ""))
        if dec == dec.to_integral_value():
            return str(int(dec))
    except (InvalidOperation, ValueError):
        pass

    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits or raw


def _build_pdf_url(doc_num: Any) -> str:
    doc = _clean_doc_number(doc_num)
    return f"{PDF_BASE_URL}/cedula_{doc}.pdf" if doc else ""


def _local_pdf_path(doc_num: Any) -> Path:
    doc = _clean_doc_number(doc_num)
    return LOCAL_PDF_DIR / f"cedula_{doc}.pdf"


def _render_pdf_pages(pdf_bytes: bytes, max_pages: int = 2, scale: float = 3.2) -> list[np.ndarray]:
    """Renderiza páginas del PDF a imágenes BGR para OCR."""
    pages_bgr: list[np.ndarray] = []
    try:
        pdf = pdfium.PdfDocument(io.BytesIO(pdf_bytes))
    except Exception:
        return pages_bgr

    page_count = min(len(pdf), max_pages)
    for i in range(page_count):
        try:
            page = pdf.get_page(i)
            bitmap = page.render(scale=scale, rotation=0)
            rgb = bitmap.to_numpy()
            page.close()
        except Exception:
            continue

        if rgb is None or rgb.size == 0:
            continue

        if len(rgb.shape) == 3 and rgb.shape[2] == 4:
            bgr = cv2.cvtColor(rgb, cv2.COLOR_RGBA2BGR)
        else:
            bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
        pages_bgr.append(bgr)

    pdf.close()
    return pages_bgr


def _extract_pdf_text(pdf_bytes: bytes, max_pages: int = 2) -> str:
    """Extrae texto nativo del PDF (cuando existe capa de texto)."""
    chunks: list[str] = []
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return ""

    page_count = min(len(reader.pages), max_pages)
    for i in range(page_count):
        try:
            text = reader.pages[i].extract_text() or ""
        except Exception:
            text = ""
        if text.strip():
            chunks.append(text)

    return "\n".join(chunks).strip()


def _save_processed_pdf_preview(tmp_dir: Path, row_idx: int, pages_bgr: list[np.ndarray], lang: str) -> str | None:
    """Crea un PDF visual con el preprocesado OCR para inspección de calidad."""
    if not pages_bgr:
        return None

    pil_pages: list[Image.Image] = []
    for page_bgr in pages_bgr:
        try:
            _, _, preview = _tutorial_link_ocr_and_preview(page_bgr, lang)
            rgb = cv2.cvtColor(preview, cv2.COLOR_BGR2RGB)
        except Exception:
            prepared = _prepare_pdf_page_for_ocr(page_bgr)
            rgb = cv2.cvtColor(prepared, cv2.COLOR_BGR2RGB)
        pil_pages.append(Image.fromarray(rgb).convert("RGB"))

    if not pil_pages:
        return None

    output_path = tmp_dir / f"row{row_idx}_procesado_{uuid.uuid4().hex}.pdf"
    first, *rest = pil_pages
    first.save(output_path, "PDF", save_all=True, append_images=rest, resolution=220.0)
    return f"/tmp_images/{output_path.name}"


def _parse_excel_date(value: Any) -> tuple[list[str], str]:
    """
    Convierte el valor de la celda I (serial, str o datetime) en:
    - lista de variantes a buscar en el OCR
    - cadena display para mostrar en la tabla
    """
    d: date | None = None

    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            d = _excel_serial_to_date(int(value))
        except Exception:
            return [], str(value)
    elif isinstance(value, str):
        v = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                d = datetime.strptime(v, fmt).date()
                break
            except ValueError:
                continue

    if d is None:
        return [], str(value) if value else ""

    display = f"{d.day:02d}/{d.month:02d}/{d.year}"
    variants = [
        f"{d.day:02d}{d.month:02d}{d.year}",
        f"{d.day:02d}/{d.month:02d}/{d.year}",
        f"{d.day}/{d.month}/{d.year}",
        f"{d.day:02d}-{d.month:02d}-{d.year}",
        str(d.year),
    ]
    return variants, display


# ── Cálculo de coincidencias ───────────────────────────────────────────────────

def calculate_match(
    ocr_text: str,
    doc_num: Any,
    nombres: list[Any],
    birth: Any,
) -> dict:
    """
    Calcula porcentaje de coincidencia entre el OCR y los 6 campos del Excel.
    Usa exact + fuzzy matching para tolerar errores tipográficos del OCR.
    """
    ocr_norm = _normalize(ocr_text)
    results: dict[str, Any] = {}
    hits = 0
    total = 0

    # Número de documento (columna D)
    found, fmt = _check_document_field(ocr_text, doc_num)
    results["documento"] = {"valor": fmt, "encontrado": found}
    total += 1
    hits += int(found)

    # Nombres y apellidos (columnas E, F, G, H)
    # Segundo nombre (F) y segundo apellido (H) son opcionales:
    # si están vacíos en el Excel se consideran válidos automáticamente.
    OPTIONAL_FIELDS = {"segundo_nombre", "segundo_apellido"}
    for label, val in zip(
        ["primer_nombre", "segundo_nombre", "primer_apellido", "segundo_apellido"],
        nombres,
    ):
        if label in OPTIONAL_FIELDS and (val is None or str(val).strip() == ""):
            results[label] = {"valor": "", "encontrado": True}
            total += 1
            hits += 1
            continue
        found, fmt = _check_name_field(ocr_norm, val)
        results[label] = {"valor": fmt, "encontrado": found}
        total += 1
        hits += int(found)

    # Fecha de nacimiento (columna I)
    date_variants, date_display = _parse_excel_date(birth)
    ocr_date_candidates = _extract_date_candidates(ocr_text)
    normalized_date_variants = {_normalize(v).replace("/", "").replace("-", "") for v in date_variants}
    date_found = any(v in ocr_date_candidates for v in normalized_date_variants if v)
    if not date_found:
        date_found = any(_normalize(v) in ocr_norm for v in date_variants)
    if not date_found and date_display:
        # Fallback fuzzy sobre la fecha completa
        ocr_words = [w for w in ocr_norm.split() if w]
        date_found = _fuzzy_word_found(ocr_words, _normalize(date_display))
    results["fecha_nacimiento"] = {"valor": date_display, "encontrado": date_found}
    total += 1
    hits += int(date_found)

    pct = round((hits / total) * 100) if total else 0
    results["porcentaje"] = pct
    results["estado"] = "OK" if pct >= 60 else "REVISAR"
    return results


# ── Procesamiento del Excel — generador SSE ────────────────────────────────────

def _get_cell(row: tuple, col: int) -> Any:
    idx = col - 1
    return row[idx] if idx < len(row) else None


def process_excel_stream(excel_path: Path, tmp_dir: Path) -> Iterator[dict]:
    """
    Generador que procesa el Excel fila a fila haciendo yield de cada resultado.
    Emite primero un evento 'meta' con el total de filas para la barra de progreso,
    luego un evento 'row' por cada fila procesada, y finalmente 'done'.
    """
    configure_tesseract()
    lang = _available_language()
    tmp_dir.mkdir(parents=True, exist_ok=True)
    LOCAL_PDF_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    COL_D, COL_E, COL_F, COL_G, COL_H, COL_I = 4, 5, 6, 7, 8, 9

    total_rows = max((ws.max_row or 1) - 1, 0)
    yield {"type": "meta", "total": total_rows}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_start = time.perf_counter()
        doc_num = _get_cell(row, COL_D)
        nombres = [
            _get_cell(row, COL_E),
            _get_cell(row, COL_F),
            _get_cell(row, COL_G),
            _get_cell(row, COL_H),
        ]
        birth = _get_cell(row, COL_I)

        # Saltar filas completamente vacías
        if not doc_num and not any(nombres):
            continue

        _, date_display = _parse_excel_date(birth)

        entry: dict[str, Any] = {
            "type": "row",
            "fila": row_idx,
            "documento": str(doc_num) if doc_num else "",
            "nombres": [str(n) if n else "" for n in nombres],
            "fecha_nacimiento": date_display or (str(birth) if birth else ""),
            "tiempo_segundos": 0.0,
            "pdf": None,
            "pdf_procesado": None,
            "imagenes": {"frontal": None, "reverso": None},
            "match": None,
            "error": None,
        }

        def trace(step: str, detail: str) -> dict:
            return {
                "type": "trace",
                "fila": row_idx,
                "documento": entry.get("documento", ""),
                "step": step,
                "detail": detail,
                "elapsed": round(time.perf_counter() - row_start, 2),
            }

        yield trace("inicio", "Iniciando validaciones de fila")

        doc = _clean_doc_number(doc_num)
        if not doc:
            entry["error"] = "Sin cédula válida para buscar PDF local"
            entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
            yield trace("validacion", "Sin cédula válida para buscar PDF local")
            yield entry
            continue

        local_pdf = _local_pdf_path(doc_num)
        entry["pdf"] = f"/pdf_cache/{local_pdf.name}"
        # Si no se requiere vista procesada distinta, usar el mismo PDF original.
        entry["pdf_procesado"] = entry["pdf"]

        yield trace("pdf", f"Buscando PDF local: {local_pdf.name}")

        combined_text = ""
        img_errors: list[str] = []

        if not local_pdf.exists() or not local_pdf.is_file():
            entry["error"] = (
                "No existe PDF local para esta cédula. "
                f"Debes guardar {local_pdf.name} en uploads/pdf_cache."
            )
            entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
            yield trace("pdf", "No existe PDF local para esta cédula")
            yield entry
            continue

        try:
            pdf_bytes = local_pdf.read_bytes()
        except Exception as exc:
            entry["error"] = f"No se pudo leer el PDF local ({local_pdf.name}): {exc}"
            entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
            yield trace("pdf", "Error al leer el PDF local")
            yield entry
            continue

        yield trace("pdf", "PDF leído correctamente")

        # 1) Primero intenta leer texto real embebido en el PDF.
        # 2) Solo lo toma como definitivo si la coincidencia preliminar es buena.
        # 3) Si no alcanza, complementa con OCR de páginas renderizadas.
        pdf_text = _extract_pdf_text(pdf_bytes, max_pages=2)
        if pdf_text and pdf_text.strip():
            yield trace("texto_pdf", "Texto embebido detectado, evaluando coincidencia preliminar")
            combined_text += " " + pdf_text
            preliminary = calculate_match(combined_text, doc_num, nombres, birth)
            # Acepta de una solo si el texto nativo ya trae buena señal.
            if preliminary.get("porcentaje", 0) >= 70:
                entry["match"] = preliminary
                entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
                yield trace("final", f"Coincidencia preliminar suficiente: {preliminary.get('porcentaje', 0)}%")
                yield entry
                continue

        # Renderiza primero solo la primera página para acelerar.
        yield trace("ocr", "Renderizando primera página para OCR")
        pages = _render_pdf_pages(pdf_bytes, max_pages=1)
        if not pages:
            if combined_text.strip():
                entry["match"] = calculate_match(combined_text, doc_num, nombres, birth)
                entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
                yield trace("final", "No hubo imágenes, usando texto disponible")
                yield entry
                continue
            entry["error"] = "No se pudo convertir el PDF a imagen"
            entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
            yield trace("ocr", "No se pudo convertir el PDF a imagen")
            yield entry
            continue

        labels = ["frontal", "reverso"]
        for idx, page_bgr in enumerate(pages):
            label = labels[idx] if idx < len(labels) else f"pagina_{idx + 1}"
            raw_path = tmp_dir / f"row{row_idx}_{label}_{uuid.uuid4().hex}_raw.jpg"
            try:
                # Guarda la página original renderizada para que el OCR pueda
                # separar mejor frontal/reverso en fondos de color.
                cv2.imwrite(str(raw_path), page_bgr)
                text, _corrected_img = preprocess_and_ocr(raw_path, lang)
                combined_text += " " + text
                yield trace("ocr", f"OCR completado en {label}")
            except Exception as exc:
                img_errors.append(str(exc))
                yield trace("ocr", f"Error OCR en {label}")
            finally:
                raw_path.unlink(missing_ok=True)

        # Si con la primera página todavía no hay buena señal, intenta la segunda.
        post_first = calculate_match(combined_text, doc_num, nombres, birth) if combined_text.strip() else {"porcentaje": 0}
        if post_first.get("porcentaje", 0) < 70:
            yield trace("ocr", "Coincidencia baja tras primera página, probando segunda página")
            extra_pages = _render_pdf_pages(pdf_bytes, max_pages=2)
            if len(extra_pages) > 1:
                page_bgr = extra_pages[1]
                raw_path = tmp_dir / f"row{row_idx}_reverso_{uuid.uuid4().hex}_raw.jpg"
                try:
                    cv2.imwrite(str(raw_path), page_bgr)
                    text, _corrected_img = preprocess_and_ocr(raw_path, lang)
                    combined_text += " " + text
                    yield trace("ocr", "OCR completado en reverso")
                except Exception as exc:
                    img_errors.append(str(exc))
                    yield trace("ocr", "Error OCR en reverso")
                finally:
                    raw_path.unlink(missing_ok=True)

        if not combined_text.strip():
            entry["error"] = "; ".join(img_errors) if img_errors else "No se extrajo texto"
            entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
            yield trace("final", "No se extrajo texto útil")
            yield entry
            continue

        entry["match"] = calculate_match(combined_text, doc_num, nombres, birth)
        if _should_mark_invalid_clarity(combined_text, entry["match"]):
            entry["match"]["estado"] = "NO VALIDA"
            entry["error"] = "No valida por claridad de imagen"
            yield trace("calidad", "Imagen marcada como no válida por claridad")
        if img_errors and not entry.get("error"):
            entry["error"] = f"Advertencia (imagen parcial): {img_errors[0]}"

        entry["tiempo_segundos"] = round(time.perf_counter() - row_start, 2)
        pct = entry.get("match", {}).get("porcentaje", 0) if entry.get("match") else 0
        yield trace("final", f"Fila completada con coincidencia {pct}%")

        yield entry

    yield {"type": "done"}
